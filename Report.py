"""
class-  Report,MeterReport,CTReport,PTReport四种类

Author: xiaohu
Date: 2022/6/8
"""
import datetime
import os
from math import modf, isclose

import openpyxl
import PySimpleGUI as sg
from docx.shared import Mm
from docxtpl import InlineImage, DocxTemplate


def pop_up(error_dict: dict, results_path, icon):
    """判断error_dict的信息并弹窗"""
    ls = list(error_dict.keys())
    vs = list(error_dict.values())
    if len(error_dict) == 1:
        sg.Popup(f'第{ls[0]}个报告有错误\n'
                 f'error:\n'
                 f'{repr(vs[0])}', icon=icon)
    elif len(error_dict) > 1:
        sg.Popup(f'第{",".join(ls)}个报告有错误\n'
                 f'error:\n'
                 f'{repr(vs)}', icon=icon)
    else:
        sg.Popup(fr'报告已生成在{results_path}', icon=icon)
    # os.startfile(results_path)


def rounding_to_str(num, interval=0.02) -> str:
    """num按照interval修约并返回修约值的字符串"""
    if type(num) not in (int, float):
        return '/'
    if type(interval) not in (int, float):
        interval = 0.02
    temp = num / interval
    tp = modf(abs(temp))
    if isclose(tp[0], 0.5, rel_tol=0.001):
        if tp[1] % 2 != 0:
            k = tp[1] + 1
        else:
            k = tp[1]
    else:
        k = round(temp)
    if temp < 0:
        result = -k * interval
    else:
        result = k * interval
    if type(interval) == float:
        n = len(str(interval)) - 2
        return f'{result:+.{n}f}'
    else:
        n = len(str(interval))
        return f'{result:+{n}d}'


class Report:
    def __init__(self, xlsx_name, docx_name, results_path: str, dic: dict):
        self.xlsx_name = xlsx_name
        self.dic = dic
        self.docx_name = docx_name
        self.docx = DocxTemplate(self.docx_name)
        self.results_path = results_path
        self.serial = dic['序号']
        self.error = {}
        self.r_number = 0
        self.interval_1 = 0.02
        self.interval_2 = 1

    @staticmethod
    def xlsx_to_dictlist(xlsx_name):
        """将xlsx文件的第一页（sheet）转化为一个字典列表并返回"""
        wb = openpyxl.load_workbook(xlsx_name, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        header = []
        dictlist = []
        for k in range(1, ws.max_column + 1):
            header.append(ws.cell(1, k).value)

        for _i in range(2, ws.max_row + 1):
            _dic = {}
            for _j in range(1, ws.max_column + 1):
                key = header[_j - 1]
                _dic[key] = ws.cell(_i, _j).value
            dictlist.append(_dic)
        return dictlist

    def xlsx_value_to_dic(self):
        """将其他页（sheet）的数据写入"""
        wb = openpyxl.load_workbook(self.xlsx_name, data_only=True, read_only=True)
        sheet = wb.get_sheet_by_name(str(self.serial))
        dic = {}
        if type(sheet['A2'].value) is int or float:
            self.interval_1 = sheet['A2'].value
        if type(sheet['A4'].value) is int or float:
            self.interval_2 = sheet['A4'].value

        x = int(sheet.max_row / 8)
        self.r_number = int(x / 3)
        for num in range(x):
            for col in 'DEFGH':
                for row in range(num * 8 + 5, num * 8 + 8, 2):
                    dic[f"{col}{row}"] = rounding_to_str(sheet[col + f'{row}'].value, self.interval_1)
                for row in range(num * 8 + 6, num * 8 + 9, 2):
                    dic[f"{col}{row}"] = rounding_to_str(sheet[col + f'{row}'].value, self.interval_2)
            for col in 'BJK':
                dic[f"{col}{num * 8 + 5}"] = sheet[f'{col}{num * 8 + 5}'].value
            dic[f"I{num * 8 + 5}"] = sheet[f'I{num * 8 + 5}'].value
            dic[f"I{num * 8 + 7}"] = sheet[f'I{num * 8 + 7}'].value
        self.dic.update(dic)

    def value_to_str(self):
        pass

    def image_deal_to_dic(self):
        """如果有照片路径处理成InlineImage对象，并更新到dic"""
        for key, value in self.dic.items():
            if "照片" in key:
                value.replace('\\', '/')
                self.dic[key] = InlineImage(self.docx, value, width=Mm(65))

    def render_to_save(self):
        try:
            self.docx.render(self.dic)
        except Exception as f:
            self.error[str(self.serial)] = repr(f)
        finally:
            self.docx.save(f'{self.results_path}/{self.serial}_{self.dic["位置"]}检验报告.docx')


class MeterReport(Report):

    def __init__(self, xlsx_name, docx_name, result_path, dic):
        super().__init__(xlsx_name, docx_name, result_path, dic)
        self.r_number = 0

    def value_to_str(self):
        """将数据格式化处理"""
        temp_dict = {}
        for key, value in self.dic.items():
            if value is None:
                self.dic[key] = r'/'
            if key in ('WD', 'W1', 'W2',
                       'W3', 'W4', '组合误差') and type(value) == float:
                n_float = self.dic['小数位']
                temp_dict[key] = f'{value:.{n_float}f}'
            elif key in ('Ua', 'Ub', 'Uc', 'Ia', 'Ib',
                         'Ic', '功率因数') and type(value) == float:
                temp_dict[key] = f'{value:.2f}'
            elif key in ('相位角A', '相位角B',
                         '相位角C', '修约误差值') and type(value) == float:
                temp_dict[key] = f'{value:+.2f}'
            elif key in ('误差1', '误差2', '平均值') and type(value) == float:
                temp_dict[key] = f'{value:+.3f}'
            elif key in ('标准时钟', '电能表时钟') and type(value) == datetime.datetime:
                temp_dict[key] = f'{value:%Y年%m月%d日 %H:%M:%S}'
            elif key == '误差' and type(value) == int:
                temp_dict[key] = f'{value:+d}'
        self.dic.update(temp_dict)


class CTReport(Report):

    def __init__(self, xlsx_name, docx_name, results_path, dic):
        super().__init__(xlsx_name, docx_name, results_path, dic)


class PTReport(Report):

    def __init__(self, xlsx_name, docx_name, results_path, dic):
        super().__init__(xlsx_name, docx_name, results_path, dic)
